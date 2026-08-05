import io
import zipfile
from pathlib import Path

import openpyxl
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase, override_settings
from rest_framework.test import APIClient

from store.models import Category, Product, ProductGalleryImage, ProductPrice, Region
from store.services import product_export
from store.services.admin_roles import ROLE_MANAGER, ROLE_MARKETING, ensure_default_admin_roles

User = get_user_model()

# A 1x1 GIF — small enough to keep the fixtures cheap, real enough for ImageField.
PIXEL_GIF = (
    b"GIF89a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00\xff\xff\xff!"
    b"\xf9\x04\x01\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;"
)


class ProductExportTestBase(TestCase):
    """Fixtures for the export suite.

    MEDIA_ROOT is redirected at a throwaway directory: these tests save real
    ImageField uploads, and the live media tree is the production catalogue.
    """

    @classmethod
    def setUpClass(cls):
        import tempfile

        cls._media_dir = tempfile.TemporaryDirectory()
        cls._media_override = override_settings(MEDIA_ROOT=cls._media_dir.name)
        cls._media_override.enable()
        super().setUpClass()

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        cls._media_override.disable()
        cls._media_dir.cleanup()

    def setUp(self):
        ensure_default_admin_roles()
        self.om = Region.objects.create(
            code="om", name_en="Oman", name_ar="عمان", currency_code="OMR",
            shipping_fee=2, shipping_threshold=0, contact_phone="1", address_en="x", address_ar="x",
        )
        self.ae = Region.objects.create(
            code="ae", name_en="UAE", name_ar="الإمارات", currency_code="AED",
            shipping_fee=2, shipping_threshold=0, contact_phone="1", address_en="x", address_ar="x",
        )
        self.category = Category.objects.create(slug="baby-care", name_en="Baby Care", name_ar="عناية")

        self.product = Product.objects.create(
            slug="lavender-baby-lotion",
            name_en="Lavender Baby Lotion",
            name_ar="غسول اللافندر للأطفال",
            unit="250 ml",
            description_en="Gentle daily lotion.",
            cost_price="1.250",
            stock_quantity=42,
            image_file=SimpleUploadedFile("main.gif", PIXEL_GIF, content_type="image/gif"),
            hover_image_file=SimpleUploadedFile("hover.gif", PIXEL_GIF, content_type="image/gif"),
            variants=[{"id": "v1", "sku": "LBL-250", "title_en": "250 ml", "price": "4.900",
                       "options": {"Size": "250 ml"}, "stock_quantity": 12}],
        )
        self.product.categories.add(self.category)
        ProductPrice.objects.create(product=self.product, region=self.om, price="4.900", compare_at_price="5.900")
        ProductPrice.objects.create(product=self.product, region=self.ae, price="46.000")
        ProductGalleryImage.objects.create(
            product=self.product,
            image_file=SimpleUploadedFile("gallery-1.gif", PIXEL_GIF, content_type="image/gif"),
            sort_order=1,
        )

        self.hidden = Product.objects.create(
            slug="unpublished-sample", name_en="Unpublished Sample", is_published=False,
        )

    def staff_client(self, role=ROLE_MANAGER, username="manager"):
        user = User.objects.create_user(username=username, password="Pass12345!", is_staff=True)
        user.groups.add(Group.objects.get(name=role))
        client = APIClient()
        client.force_authenticate(user)
        return client


class ProductImagePlanTests(ProductExportTestBase):
    """Every stored image shape has to resolve to a real file, exactly once."""

    def test_plans_cover_main_hover_and_gallery_without_duplicates(self):
        plans = product_export.plan_product_images(self.product)
        roles = [plan.role for plan in plans]
        self.assertEqual(roles.count(product_export.ROLE_MAIN), 1)
        self.assertEqual(roles.count(product_export.ROLE_HOVER), 1)
        self.assertEqual(roles.count(product_export.ROLE_GALLERY), 1)
        self.assertTrue(all(plan.kind == "local" for plan in plans), [p.kind for p in plans])

    def test_gallery_json_pointing_at_the_same_file_is_not_exported_twice(self):
        # The admin gallery widget stores /media/... URLs that repeat the main image.
        self.product.gallery = [self.product.image_file.url, self.product.image_file.url]
        self.product.save(update_fields=["gallery"])
        plans = product_export.plan_product_images(self.product)
        local_paths = [str(plan.local_path) for plan in plans if plan.local_path]
        self.assertEqual(len(local_paths), len(set(local_paths)))

    def test_remote_url_is_marked_remote_and_can_be_skipped(self):
        self.product.gallery = ["https://cdn.shopify.com/s/files/example.png"]
        self.product.save(update_fields=["gallery"])

        remote = [p for p in product_export.plan_product_images(self.product) if p.source.startswith("https://")]
        self.assertEqual(len(remote), 1)
        self.assertEqual(remote[0].kind, "remote")

        skipped = [
            p for p in product_export.plan_product_images(self.product, include_remote=False)
            if p.source.startswith("https://")
        ]
        self.assertEqual(skipped[0].kind, "skipped")

    def test_missing_file_is_reported_rather_than_crashing_the_export(self):
        self.product.gallery = ["products/does-not-exist.jpg"]
        self.product.save(update_fields=["gallery"])
        plans = product_export.plan_product_images(self.product)
        missing = [p for p in plans if p.kind == "missing"]
        self.assertEqual(len(missing), 1)

    def test_each_product_gets_its_own_folder_and_unique_file_names(self):
        products = list(product_export.export_queryset())
        plans = product_export.build_plans(products)
        folders = {
            plan.arcname.split("/")[1]
            for product_plans in plans.values() for plan in product_plans
        }
        self.assertTrue(all(folder.startswith(("001-", "002-")) for folder in folders), folders)
        for product_plans in plans.values():
            names = [plan.file_name for plan in product_plans]
            self.assertEqual(len(names), len(set(names)))


class ProductWorkbookTests(ProductExportTestBase):
    def test_workbook_carries_every_sheet_and_per_region_prices(self):
        products = list(product_export.export_queryset())
        plans = product_export.build_plans(products)
        workbook = product_export.build_workbook(products, plans)

        self.assertEqual(
            workbook.sheetnames,
            ["Summary", "Products", "Prices", "Stock", "Variants", "Images", "Categories", "Reviews"],
        )

        sheet = workbook["Products"]
        header = [cell.value for cell in sheet[1]]
        self.assertIn("Price OM (OMR)", header)
        self.assertIn("Price AE (AED)", header)
        self.assertIn("Image Folder", header)

        rows = {row[header.index("Slug")]: row for row in sheet.iter_rows(min_row=2, values_only=True)}
        row = rows["lavender-baby-lotion"]
        self.assertEqual(row[header.index("Name (AR)")], "غسول اللافندر للأطفال")
        self.assertEqual(row[header.index("Price OM (OMR)")], 4.9)
        self.assertEqual(row[header.index("Price AE (AED)")], 46.0)
        self.assertEqual(row[header.index("Stock Quantity")], 42)
        self.assertEqual(row[header.index("Published")], "Yes")
        self.assertEqual(rows["unpublished-sample"][header.index("Published")], "No")

        # The folder named in Excel is the folder the images actually land in.
        folder = row[header.index("Image Folder")]
        arcnames = {plan.arcname for plan in plans[self.product.pk]}
        self.assertTrue(all(name.startswith(f"images/{folder}/") for name in arcnames), arcnames)

    def test_variant_and_price_sheets_are_populated(self):
        products = list(product_export.export_queryset())
        plans = product_export.build_plans(products)
        workbook = product_export.build_workbook(products, plans)

        variant_rows = list(workbook["Variants"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(variant_rows), 1)
        self.assertEqual(variant_rows[0][3], "LBL-250")
        self.assertEqual(variant_rows[0][6], "Size: 250 ml")

        price_rows = list(workbook["Prices"].iter_rows(min_row=2, values_only=True))
        self.assertEqual({row[2] for row in price_rows}, {"OM", "AE"})

    def test_published_only_drops_hidden_products(self):
        slugs = [p.slug for p in product_export.export_queryset(include_unpublished=False)]
        self.assertNotIn("unpublished-sample", slugs)

    def test_leading_equals_is_not_written_as_a_formula(self):
        self.product.short_description_en = "=SUM(A1:A2)"
        self.product.save(update_fields=["short_description_en"])
        products = list(product_export.export_queryset())
        workbook = product_export.build_workbook(products, product_export.build_plans(products, include_images=False))
        sheet = workbook["Products"]
        header = [cell.value for cell in sheet[1]]
        row = next(r for r in sheet.iter_rows(min_row=2, values_only=True) if r[header.index("Slug")] == self.product.slug)
        self.assertEqual(row[header.index("Short Description (EN)")], "'=SUM(A1:A2)")


class ProductExportZipTests(ProductExportTestBase):
    def test_zip_stream_contains_workbook_readme_and_per_product_image_folders(self):
        products = list(product_export.export_queryset())
        plans = product_export.build_plans(products)
        result = product_export.ExportResult()
        buffer = io.BytesIO()
        for chunk in product_export.iter_export_zip(
            products, plans, root_name="catalogue", result=result
        ):
            buffer.write(chunk)

        buffer.seek(0)
        archive = zipfile.ZipFile(buffer)
        self.assertIsNone(archive.testzip())
        names = archive.namelist()
        self.assertIn("catalogue/products.xlsx", names)
        self.assertIn("catalogue/README.txt", names)

        image_names = [n for n in names if n.startswith("catalogue/images/")]
        self.assertEqual(len(image_names), 3)
        self.assertEqual(result.images_exported, 3)
        self.assertEqual(result.images_failed, 0)
        for name in image_names:
            self.assertEqual(archive.read(name), PIXEL_GIF)

        # The workbook inside the zip must agree with what was actually written.
        workbook = openpyxl.load_workbook(io.BytesIO(archive.read("catalogue/products.xlsx")))
        statuses = {row[5] for row in workbook["Images"].iter_rows(min_row=2, values_only=True)}
        self.assertEqual(statuses, {"exported"})


class ProductExportApiTests(ProductExportTestBase):
    def test_xlsx_export_returns_a_workbook(self):
        response = self.staff_client().get("/api/admin/products/export/?images=0")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertEqual(response["X-Export-Products"], "2")
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn("Products", workbook.sheetnames)

    def test_zip_export_streams_images(self):
        response = self.staff_client().get("/api/admin/products/export/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertEqual(response["X-Accel-Buffering"], "no")
        archive = zipfile.ZipFile(io.BytesIO(b"".join(response.streaming_content)))
        self.assertIsNone(archive.testzip())
        self.assertEqual(len([n for n in archive.namelist() if "/images/" in n]), 3)

    def test_published_only_flag_is_honoured(self):
        response = self.staff_client().get("/api/admin/products/export/?images=0&published_only=1")
        self.assertEqual(response["X-Export-Products"], "1")

    def test_export_requires_authentication(self):
        self.assertEqual(APIClient().get("/api/admin/products/export/?images=0").status_code, 401)

    def test_role_without_products_capability_is_refused(self):
        client = self.staff_client(role=ROLE_MARKETING, username="marketer")
        self.assertEqual(client.get("/api/admin/products/export/?images=0").status_code, 403)

    def test_export_is_written_to_the_audit_log(self):
        from store.models import AdminAuditLog

        self.staff_client().get("/api/admin/products/export/?images=0")
        entry = AdminAuditLog.objects.filter(resource_type="product", resource_id="catalogue-export").first()
        self.assertIsNotNone(entry)
        self.assertEqual(entry.action, "export")
        self.assertEqual(entry.after_snapshot["format"], "xlsx")


class ExportProductsCommandTests(ProductExportTestBase):
    def test_command_writes_folder_with_workbook_readme_and_images(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            out = io.StringIO()
            call_command("export_products", "--output", tmp, "--name", "catalogue", stdout=out)
            root = Path(tmp) / "catalogue"

            self.assertTrue((root / "products.xlsx").is_file())
            self.assertTrue((root / "README.txt").is_file())

            folders = sorted(p.name for p in (root / "images").iterdir())
            self.assertEqual(folders, ["001-lavender-baby-lotion"])
            files = sorted(p.name for p in (root / "images" / folders[0]).iterdir())
            self.assertEqual(len(files), 3)
            self.assertTrue(files[0].startswith("01-main-"))
            self.assertTrue(files[1].startswith("02-hover-"))
            self.assertTrue(files[2].startswith("03-gallery-"))

    def test_no_images_flag_writes_the_workbook_only(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            call_command("export_products", "--output", tmp, "--name", "data-only",
                         "--no-images", stdout=io.StringIO())
            root = Path(tmp) / "data-only"
            self.assertTrue((root / "products.xlsx").is_file())
            self.assertFalse((root / "images").exists())

    def test_zip_flag_produces_an_archive_next_to_the_folder(self):
        import tempfile

        with tempfile.TemporaryDirectory() as tmp:
            call_command("export_products", "--output", tmp, "--name", "catalogue",
                         "--zip", stdout=io.StringIO())
            archive_path = Path(tmp) / "catalogue.zip"
            self.assertTrue(archive_path.is_file())
            archive = zipfile.ZipFile(archive_path)
            self.assertIn("catalogue/products.xlsx", archive.namelist())

    def test_rerunning_without_force_refuses_to_overwrite(self):
        import tempfile

        from django.core.management.base import CommandError

        with tempfile.TemporaryDirectory() as tmp:
            call_command("export_products", "--output", tmp, "--name", "catalogue",
                         "--no-images", stdout=io.StringIO())
            with self.assertRaises(CommandError):
                call_command("export_products", "--output", tmp, "--name", "catalogue",
                             "--no-images", stdout=io.StringIO())
            call_command("export_products", "--output", tmp, "--name", "catalogue",
                         "--no-images", "--force", stdout=io.StringIO())
