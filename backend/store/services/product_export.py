"""Catalogue export — one Excel workbook plus per-product image folders.

The deliverable the client asked for looks like this::

    enfant-products-export-2026-08-05/
    ├── products.xlsx          all product data, one sheet per concern
    ├── README.txt             what is inside + how the folders map to rows
    └── images/
        ├── 001-lavender-baby-lotion/
        │   ├── 01-main-p16-primary.jpeg
        │   └── 02-hover-p16-hover.jpeg
        └── 002-…/

The same builder backs three entry points so the structure never drifts:

* ``python manage.py export_products`` writes the folder (or a zip) to disk,
* ``GET /api/admin/products/export/`` streams the zip to the admin browser,
* ``GET /api/admin/products/export/?images=0`` streams only the workbook.

Images are streamed into the archive one at a time and never all held in
memory — the media tree runs to hundreds of megabytes, so anything that buffers
the whole export would take the web worker down with it.
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
import shutil
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from urllib.parse import unquote, urlparse

from django.conf import settings
from django.utils import timezone
from django.utils.text import slugify

from ..models import Category, Product, Region, Review

logger = logging.getLogger(__name__)

# Excel refuses to open a file with a cell longer than this.
EXCEL_CELL_LIMIT = 32_767
# Remote images bigger than this are skipped rather than stalling the export.
MAX_REMOTE_IMAGE_BYTES = 25 * 1024 * 1024
REMOTE_TIMEOUT_SECONDS = 20
# How much zip output to accumulate before handing a chunk to the WSGI server.
STREAM_CHUNK_BYTES = 512 * 1024
FILE_COPY_CHUNK_BYTES = 256 * 1024
# Folder names stay short so the deep paths survive Windows' 260-char limit.
MAX_FOLDER_SLUG_CHARS = 60

ROLE_MAIN = "main"
ROLE_HOVER = "hover"
ROLE_GALLERY = "gallery"
ROLE_VARIANT = "variant"
ROLE_CERTIFICATE = "certificate"

_UNSAFE_FILENAME_RE = re.compile(r"[^A-Za-z0-9._-]+")
# (path, size, mtime) -> md5, so a file shared by several products is only read
# once. Keyed on stat as well as path: a worker process outlives many exports,
# and media replaced in place under an existing name must not return a stale
# digest.
_DIGEST_CACHE: dict = {}


# ---------------------------------------------------------------------------
# value helpers
# ---------------------------------------------------------------------------

def _cell(value):
    """Coerce any model value into something openpyxl will accept."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return timezone.localtime(value).strftime("%Y-%m-%d %H:%M") if timezone.is_aware(value) else value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (list, dict)):
        value = json.dumps(value, ensure_ascii=False)
    if isinstance(value, str):
        # openpyxl turns a leading "=" into a live formula; product copy is text.
        if value.startswith("="):
            value = "'" + value
        if len(value) > EXCEL_CELL_LIMIT:
            return value[: EXCEL_CELL_LIMIT - 20] + "… [truncated]"
        return value
    return value


def _join(values, separator=" | "):
    return separator.join(str(v) for v in values if str(v or "").strip())


def _safe_filename(name, fallback="image"):
    name = unquote(str(name or "")).strip().replace("\\", "/").split("/")[-1]
    name = _UNSAFE_FILENAME_RE.sub("-", name).strip("-._")
    if not name:
        name = fallback
    stem, dot, suffix = name.rpartition(".")
    if dot and len(suffix) <= 5:
        return f"{stem[:60]}.{suffix.lower()}"
    return name[:64]


def _product_folder(index, product):
    slug = slugify(product.slug or product.name_en or f"product-{product.pk}")[:MAX_FOLDER_SLUG_CHARS]
    return f"{index:03d}-{slug or f'product-{product.pk}'}"


def _storefront_url(product, locale="en", region_code="om"):
    base = (
        getattr(settings, "MEDIA_HOST_URL", "")
        or getattr(settings, "FRONTEND_PUBLIC_URL", "")
        or ""
    ).rstrip("/")
    if not base:
        return ""
    return f"{base}/{locale}-{region_code}/product/{product.slug}"


# ---------------------------------------------------------------------------
# image resolution
# ---------------------------------------------------------------------------

@dataclass
class ImagePlan:
    """One image we intend to place inside a product's folder."""

    role: str
    source: str          # original value as stored on the product
    kind: str            # "local" | "remote" | "missing"
    local_path: Path | None = None
    remote_url: str = ""
    arcname: str = ""    # filled in once the folder is known
    file_name: str = ""
    status: str = ""     # filled in after the bytes are written
    size_bytes: int = 0
    note: str = ""


def _media_root() -> Path:
    return Path(str(settings.MEDIA_ROOT))


def _strip_media_prefix(path: str) -> str | None:
    media_url = str(getattr(settings, "MEDIA_URL", "/media/") or "/media/")
    if media_url and path.startswith(media_url):
        return path[len(media_url):]
    if path.startswith("/media/"):
        return path[len("/media/"):]
    return None


def _local_media_file(relative) -> Path | None:
    """Resolve a media-relative path, refusing anything that escapes MEDIA_ROOT.

    These values are admin-editable, so a ``../..`` in a gallery entry must not
    be able to pull arbitrary server files into a customer-facing export.
    """
    root = _media_root().resolve()
    try:
        candidate = (root / str(relative)).resolve()
    except (OSError, ValueError):
        return None
    if candidate != root and root not in candidate.parents:
        return None
    return candidate if candidate.is_file() else None


def _resolve_source(raw) -> ImagePlan | None:
    """Turn a stored image reference into a local path or a remote URL.

    Handles every shape the catalogue actually uses: ``ImageField`` files,
    ``/media/...`` URLs, absolute URLs pointing back at our own domain, bare
    storage-relative paths, and third-party CDN links left over from Shopify.

    A ``/media`` path that exists on disk is always taken from disk, whatever
    host the URL names. Products routinely store the same file both as an
    ``ImageField`` and as an absolute URL, and matching the URL's host against
    our configured domains got that wrong whenever the two disagreed (the
    backend knows itself as ``enfhantorganic.itwing.cloud`` while the images are
    written as ``www.enfantorganic.com``) — the file was then re-downloaded over
    the network and landed in the export twice.
    """
    if raw is None:
        return None

    # ImageField / FileField
    if hasattr(raw, "name") and hasattr(raw, "storage"):
        if not raw.name:
            return None
        candidate = _local_media_file(raw.name)
        if candidate:
            return ImagePlan(role="", source=str(raw.name), kind="local", local_path=candidate)
        try:
            url = raw.url
        except Exception:  # storage without a public URL
            url = ""
        if url:
            return _resolve_source(url)
        return ImagePlan(role="", source=str(raw.name), kind="missing", note="file not found on disk")

    value = str(raw or "").strip()
    if not value:
        return None

    if value.startswith("//"):
        value = "https:" + value

    if value.startswith("http://") or value.startswith("https://"):
        relative = _strip_media_prefix(unquote(urlparse(value).path))
        if relative is not None:
            candidate = _local_media_file(relative)
            if candidate:
                return ImagePlan(role="", source=value, kind="local", local_path=candidate)
        return ImagePlan(role="", source=value, kind="remote", remote_url=value)

    relative = _strip_media_prefix(unquote(value))
    if relative is None:
        relative = unquote(value).lstrip("/")
    candidate = _local_media_file(relative)
    if candidate:
        return ImagePlan(role="", source=value, kind="local", local_path=candidate)
    return ImagePlan(role="", source=value, kind="missing", note="file not found on disk")


def _gallery_entries(product):
    for entry in product.gallery if isinstance(product.gallery, list) else []:
        if isinstance(entry, dict):
            yield entry.get("url") or entry.get("image") or entry.get("src")
        else:
            yield entry


def _variant_images(product):
    for variant in product.variants if isinstance(product.variants, list) else []:
        if isinstance(variant, dict) and variant.get("image"):
            yield variant["image"]


def _content_digest(path: Path) -> str:
    """MD5 of a media file, cached per path for the life of the process.

    Not a security hash — it only answers "are these two files the same
    picture?", which the storage paths cannot: the admin gallery widget saves
    its own copy under a fresh random name, so one photo can sit on disk two or
    three times under names that share nothing.
    """
    stat = path.stat()
    key = (str(path), stat.st_size, stat.st_mtime_ns)
    cached = _DIGEST_CACHE.get(key)
    if cached is not None:
        return cached
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(FILE_COPY_CHUNK_BYTES), b""):
            digest.update(block)
    value = digest.hexdigest()
    _DIGEST_CACHE[key] = value
    return value


def plan_product_images(product, *, include_remote=True) -> list[ImagePlan]:
    """Every distinct image attached to ``product``, in display order.

    De-duplication runs on file content, not on the stored reference, and is
    scoped to the product — the same photo shared by two products belongs in
    both folders, but it must not appear twice inside one.
    """
    raw_sources = [(ROLE_MAIN, product.image_file), (ROLE_MAIN, product.image),
                   (ROLE_HOVER, product.hover_image_file), (ROLE_HOVER, product.hover_image)]
    for row in product.gallery_images.all():
        raw_sources.append((ROLE_GALLERY, row.image_file))
        raw_sources.append((ROLE_GALLERY, row.image_url))
    for entry in _gallery_entries(product):
        raw_sources.append((ROLE_GALLERY, entry))
    for entry in _variant_images(product):
        raw_sources.append((ROLE_VARIANT, entry))
    raw_sources.append((ROLE_CERTIFICATE, product.organic_certification_file))

    plans: list[ImagePlan] = []
    seen: set[str] = set()
    for role, raw in raw_sources:
        plan = _resolve_source(raw)
        if plan is None:
            continue
        if plan.kind == "local" and plan.local_path:
            try:
                key = f"md5:{_content_digest(plan.local_path)}"
            except OSError:
                key = str(plan.local_path)
        else:
            key = plan.remote_url or plan.source
        if key in seen:
            continue
        seen.add(key)
        if plan.kind == "remote" and not include_remote:
            plan.kind = "skipped"
            plan.note = "remote download disabled"
        plan.role = role
        plans.append(plan)
    return plans


# ---------------------------------------------------------------------------
# workbook
# ---------------------------------------------------------------------------

PRODUCT_TEXT_FIELDS = [
    ("short_description_en", "Short Description (EN)"),
    ("short_description_ar", "Short Description (AR)"),
    ("description_en", "Description (EN)"),
    ("description_ar", "Description (AR)"),
    ("ingredients_en", "Ingredients (EN)"),
    ("ingredients_ar", "Ingredients (AR)"),
    ("usage_instructions_en", "Usage Instructions (EN)"),
    ("usage_instructions_ar", "Usage Instructions (AR)"),
    ("origin_source_en", "Origin / Source (EN)"),
    ("origin_source_ar", "Origin / Source (AR)"),
]

SEO_FIELDS = [
    ("seo_title_en", "SEO Title (EN)"),
    ("seo_title_ar", "SEO Title (AR)"),
    ("seo_description_en", "SEO Description (EN)"),
    ("seo_description_ar", "SEO Description (AR)"),
    ("canonical_url", "Canonical URL"),
    ("og_title_en", "OG Title (EN)"),
    ("og_title_ar", "OG Title (AR)"),
    ("og_description_en", "OG Description (EN)"),
    ("og_description_ar", "OG Description (AR)"),
    ("og_image", "OG Image"),
]


def _autosize(sheet, headers, sample_rows, max_width=52):
    from openpyxl.utils import get_column_letter

    widths = [len(str(h)) + 4 for h in headers]
    for row in sample_rows:
        for idx, value in enumerate(row):
            if idx >= len(widths):
                break
            widths[idx] = max(widths[idx], min(len(str(value)) + 2, max_width))
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(width, 10), max_width)


def _write_sheet(workbook, title, headers, rows, *, freeze="A2"):
    from openpyxl.styles import Alignment, Font, PatternFill

    sheet = workbook.create_sheet(title=title)
    sheet.append(headers)
    for row in rows:
        sheet.append([_cell(value) for value in row])

    header_fill = PatternFill("solid", fgColor="1F3A2E")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    if freeze:
        sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    _autosize(sheet, headers, rows[:60])
    return sheet


def build_workbook(products, plans_by_pk, *, regions=None, include_reviews=True, options=None):
    """Assemble the multi-sheet workbook for ``products``.

    ``plans_by_pk`` maps product id → list of :class:`ImagePlan` so the Images
    sheet can state exactly which file landed in which folder (and which ones
    could not be fetched).
    """
    import openpyxl

    regions = list(regions if regions is not None else Region.objects.filter(is_active=True).order_by("sort_order", "id"))
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # -- Products -----------------------------------------------------------
    headers = [
        "#", "Image Folder", "Slug", "Name (EN)", "Name (AR)", "Brand", "Unit",
        "Vendor (EN)", "Vendor (AR)", "Categories", "Tags", "Published", "Featured",
        "New Arrivals", "Baby Sets", "Top Choices", "Badge (EN)", "Badge (AR)",
        "Rating", "Review Count", "Cost Price",
    ]
    for region in regions:
        headers.append(f"Price {region.code.upper()} ({region.currency_code})")
        headers.append(f"Compare At {region.code.upper()}")
    headers += [
        "Stock Quantity", "Track Inventory", "Warehouse Qty", "Warehouse Available",
        "Images In Export", "Missing Images", "Main Image", "Hover Image", "Gallery Count",
        "Dietary Tags", "Shelf Life", "Expiry Date", "Organic Certification",
        "Option Groups (EN)", "Option Groups (AR)", "Variant Count",
        "Details (EN)", "Details (AR)",
    ]
    headers += [label for _, label in PRODUCT_TEXT_FIELDS]
    headers += [label for _, label in SEO_FIELDS]
    headers += ["Meta Robots Index", "Meta Robots Follow", "Storefront URL", "Sort Order", "Product ID"]

    product_rows = []
    price_rows = []
    stock_rows = []
    variant_rows = []
    image_rows = []

    for index, product in enumerate(products, start=1):
        plans = plans_by_pk.get(product.pk, [])
        folder = _product_folder(index, product)
        price_map = {price.region_id: price for price in product.prices.all()}
        stocks = list(product.warehouse_stocks.all())
        warehouse_qty = sum(int(s.quantity or 0) for s in stocks)
        warehouse_available = sum(s.available_quantity for s in stocks)
        variants = product.variants if isinstance(product.variants, list) else []
        exported_images = [p for p in plans if p.status == "exported"]
        failed_images = [p for p in plans if p.status not in {"exported", ""}]

        row = [
            index, folder, product.slug, product.name_en, product.name_ar, product.brand,
            product.unit, product.vendor_en, product.vendor_ar,
            _join(c.name_en or c.slug for c in product.categories.all()),
            _join(t.name_en if hasattr(t, "name_en") else str(t) for t in product.tags.all()),
            product.is_published, product.is_featured, product.show_in_new_arrivals,
            product.show_in_baby_sets, product.show_in_top_choices,
            product.badge_en, product.badge_ar, product.rating, product.review_count,
            product.cost_price,
        ]
        for region in regions:
            price = price_map.get(region.id)
            row.append(price.price if price else "")
            row.append(price.compare_at_price if price and price.compare_at_price is not None else "")
        row += [
            product.stock_quantity, product.track_inventory, warehouse_qty, warehouse_available,
            len(exported_images), len(failed_images),
            product.image or (product.image_file.name if product.image_file else ""),
            product.hover_image or (product.hover_image_file.name if product.hover_image_file else ""),
            len([p for p in plans if p.role == ROLE_GALLERY]),
            product.dietary_tags, product.shelf_life, product.expiry_date,
            product.organic_certification_name,
            product.option_groups_en, product.option_groups_ar, len(variants),
            product.details_en, product.details_ar,
        ]
        row += [getattr(product, name, "") for name, _ in PRODUCT_TEXT_FIELDS]
        row += [getattr(product, name, "") for name, _ in SEO_FIELDS]
        row += [
            product.meta_robots_index, product.meta_robots_follow,
            _storefront_url(product), product.sort_order, product.pk,
        ]
        product_rows.append(row)

        for region in regions:
            price = price_map.get(region.id)
            if not price:
                continue
            price_rows.append([
                product.slug, product.name_en, region.code.upper(), region.currency_code,
                price.price, price.compare_at_price, price.price_prefix_en, price.price_prefix_ar,
                price.unit_price_text_en, price.unit_price_text_ar,
            ])

        for stock in stocks:
            stock_rows.append([
                product.slug, product.name_en, stock.warehouse.code.upper(),
                stock.warehouse.name_en, stock.warehouse.region.code.upper(),
                stock.quantity, stock.reserved_quantity, stock.available_quantity,
                stock.low_stock_threshold, stock.is_low_stock,
            ])

        for v_index, variant in enumerate(variants, start=1):
            if not isinstance(variant, dict):
                continue
            options_value = variant.get("options") if isinstance(variant.get("options"), dict) else {}
            variant_rows.append([
                product.slug, product.name_en,
                variant.get("id") or f"variant-{v_index}", variant.get("sku", ""),
                variant.get("title_en", ""), variant.get("title_ar", ""),
                _join(f"{k}: {v}" for k, v in options_value.items()),
                variant.get("price", variant.get("base_price", "")),
                variant.get("compare_at_price", variant.get("base_compare_at_price", "")),
                variant.get("stock_quantity", ""),
                variant.get("is_active", True),
                variant.get("image", ""),
            ])

        for plan in plans:
            image_rows.append([
                product.slug, product.name_en, folder, plan.file_name, plan.role,
                plan.status or plan.kind, plan.size_bytes or "", plan.source, plan.note,
            ])

    _write_sheet(workbook, "Products", headers, product_rows)
    _write_sheet(workbook, "Prices", [
        "Slug", "Product", "Region", "Currency", "Price", "Compare At Price",
        "Price Prefix (EN)", "Price Prefix (AR)", "Unit Price Text (EN)", "Unit Price Text (AR)",
    ], price_rows)
    _write_sheet(workbook, "Stock", [
        "Slug", "Product", "Warehouse", "Warehouse Name", "Region",
        "Quantity", "Reserved", "Available", "Low Stock Threshold", "Is Low Stock",
    ], stock_rows)
    _write_sheet(workbook, "Variants", [
        "Slug", "Product", "Variant ID", "SKU", "Title (EN)", "Title (AR)",
        "Options", "Price", "Compare At Price", "Stock Quantity", "Active", "Image",
    ], variant_rows)
    _write_sheet(workbook, "Images", [
        "Slug", "Product", "Folder", "File Name", "Role", "Status", "Size (bytes)",
        "Original Source", "Note",
    ], image_rows)

    category_rows = [
        [c.slug, c.name_en, c.name_ar, c.description_en, c.category_products.count(), c.sort_order]
        for c in Category.objects.all().order_by("sort_order", "id")
    ]
    _write_sheet(workbook, "Categories", [
        "Slug", "Name (EN)", "Name (AR)", "Description (EN)", "Product Count", "Sort Order",
    ], category_rows)

    if include_reviews:
        product_ids = [p.pk for p in products]
        review_rows = [
            [
                review.product.slug, review.product.name_en, review.customer_name, review.rating,
                review.title, review.comment, review.is_verified_purchase, review.is_approved,
                review.created_at,
            ]
            for review in Review.objects.filter(product_id__in=product_ids)
            .select_related("product").order_by("product__slug", "-created_at")[:20000]
        ]
        _write_sheet(workbook, "Reviews", [
            "Slug", "Product", "Customer", "Rating", "Title", "Comment",
            "Verified Purchase", "Approved", "Created At",
        ], review_rows)

    # -- Summary (first sheet) ---------------------------------------------
    total_images = sum(len(v) for v in plans_by_pk.values())
    exported = sum(1 for v in plans_by_pk.values() for p in v if p.status == "exported")
    summary_rows = [
        ["Generated at", timezone.localtime().strftime("%Y-%m-%d %H:%M:%S %Z")],
        ["Products exported", len(products)],
        ["Published products", sum(1 for p in products if p.is_published)],
        ["Unpublished products", sum(1 for p in products if not p.is_published)],
        ["Regions with prices", _join(r.code.upper() for r in regions)],
        ["Price rows", len(price_rows)],
        ["Warehouse stock rows", len(stock_rows)],
        ["Variant rows", len(variant_rows)],
        ["Images referenced", total_images],
        ["Images exported", exported],
        ["Images missing / failed", total_images - exported],
        ["Categories", len(category_rows)],
    ]
    for key, value in (options or {}).items():
        summary_rows.append([f"Option · {key}", value])
    summary = _write_sheet(workbook, "Summary", ["Field", "Value"], summary_rows, freeze=None)
    summary.auto_filter.ref = None
    workbook.move_sheet("Summary", offset=-(len(workbook.sheetnames) - 1))
    return workbook


def workbook_bytes(workbook) -> bytes:
    import io

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# export driver
# ---------------------------------------------------------------------------

def export_queryset(*, include_unpublished=True):
    queryset = (
        Product.objects.all()
        .prefetch_related(
            "categories", "tags", "gallery_images",
            "prices__region", "warehouse_stocks__warehouse__region",
        )
        .order_by("sort_order", "id")
    )
    if not include_unpublished:
        queryset = queryset.filter(is_published=True)
    return queryset


def export_base_name(prefix="enfant-products-export"):
    return f"{prefix}-{timezone.localdate().isoformat()}"


def _readme_text(product_count, options):
    lines = [
        "ENFANT ORGANIC — PRODUCT CATALOGUE EXPORT",
        "=" * 46,
        "",
        f"Generated : {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S %Z')}",
        f"Products  : {product_count}",
        "",
        "CONTENTS",
        "--------",
        "products.xlsx   All product data. Sheets:",
        "                  Summary    – export counts and settings",
        "                  Products   – one row per product (every field, prices per region)",
        "                  Prices     – price / compare-at price per region",
        "                  Stock      – quantity per warehouse",
        "                  Variants   – variant rows where a product has them",
        "                  Images     – every image file, which folder it is in, and its source",
        "                  Categories – category list with product counts",
        "                  Reviews    – customer reviews per product",
        "",
        "images/         One folder per product, named  <row number>-<product slug>.",
        "                The same folder name is in the 'Image Folder' column of the",
        "                Products sheet, so a row in Excel maps straight to a folder.",
        "                Files are named  <order>-<role>-<original file name>, where role",
        "                is main / hover / gallery / variant / certificate.",
        "",
        "NOTES",
        "-----",
        "* Images are exported at their original uploaded resolution.",
        "* A product with no images gets no folder; check the 'Images In Export'",
        "  and 'Missing Images' columns in the Products sheet.",
        "* Prices are stored per region; the region columns in the Products sheet",
        "  carry each market's own currency (OMR / AED / SAR).",
    ]
    if options:
        lines += ["", "EXPORT OPTIONS", "--------------"]
        lines += [f"* {key}: {value}" for key, value in options.items()]
    return "\n".join(lines) + "\n"


class _StreamBuffer:
    """Write-only sink that ``zipfile`` can target on an unseekable stream."""

    def __init__(self):
        self._chunks = bytearray()
        self._position = 0

    def write(self, data):
        self._chunks.extend(data)
        self._position += len(data)
        return len(data)

    def tell(self):
        return self._position

    def flush(self):
        return None

    def pending(self):
        return len(self._chunks)

    def take(self):
        chunk = bytes(self._chunks)
        self._chunks.clear()
        return chunk


@dataclass
class ExportResult:
    products: int = 0
    images_exported: int = 0
    images_failed: int = 0
    bytes_written: int = 0
    failures: list = field(default_factory=list)


def _fetch_remote(url):
    import requests

    response = requests.get(url, timeout=REMOTE_TIMEOUT_SECONDS, stream=True)
    response.raise_for_status()
    declared = response.headers.get("Content-Length")
    if declared and int(declared) > MAX_REMOTE_IMAGE_BYTES:
        raise ValueError(f"image is {int(declared)} bytes, over the {MAX_REMOTE_IMAGE_BYTES} byte cap")
    payload = bytearray()
    for chunk in response.iter_content(FILE_COPY_CHUNK_BYTES):
        payload.extend(chunk)
        if len(payload) > MAX_REMOTE_IMAGE_BYTES:
            raise ValueError(f"image exceeds the {MAX_REMOTE_IMAGE_BYTES} byte cap")
    return bytes(payload)


def assign_arcnames(products, plans_by_pk):
    """Give every plan its final folder + file name before anything is written."""
    for index, product in enumerate(products, start=1):
        folder = _product_folder(index, product)
        used = set()
        for order, plan in enumerate(plans_by_pk.get(product.pk, []), start=1):
            base = plan.local_path.name if plan.local_path else (plan.remote_url or plan.source)
            name = f"{order:02d}-{plan.role}-{_safe_filename(base)}"
            while name.lower() in used:
                name = f"{order:02d}-{plan.role}-{len(used)}-{_safe_filename(base)}"
            used.add(name.lower())
            plan.file_name = name
            plan.arcname = f"images/{folder}/{name}"


def build_plans(products, *, include_images=True, include_remote=True):
    plans_by_pk = {}
    for product in products:
        plans_by_pk[product.pk] = plan_product_images(product, include_remote=include_remote) if include_images else []
    assign_arcnames(products, plans_by_pk)
    return plans_by_pk


def iter_export_zip(
    products,
    plans_by_pk,
    *,
    root_name,
    include_images=True,
    result=None,
    options=None,
):
    """Yield the export archive chunk by chunk.

    Images go in first so bytes start flowing immediately (nginx will not sit
    on an idle connection), and ``products.xlsx`` is written last — by then
    every image has a real status to record in the Images sheet.
    """
    result = result if result is not None else ExportResult()
    result.products = len(products)
    buffer = _StreamBuffer()

    with zipfile.ZipFile(buffer, "w", allowZip64=True) as archive:
        if include_images:
            for product in products:
                for plan in plans_by_pk.get(product.pk, []):
                    arcname = f"{root_name}/{plan.arcname}"
                    try:
                        if plan.kind == "local" and plan.local_path and plan.local_path.is_file():
                            plan.size_bytes = plan.local_path.stat().st_size
                            info = zipfile.ZipInfo(arcname, date_time=timezone.localtime().timetuple()[:6])
                            info.compress_type = zipfile.ZIP_STORED
                            # Without this the entries carry mode 0, and some
                            # extractors then produce unreadable files.
                            info.external_attr = 0o644 << 16
                            with archive.open(info, "w") as target, plan.local_path.open("rb") as source:
                                shutil.copyfileobj(source, target, FILE_COPY_CHUNK_BYTES)
                            plan.status = "exported"
                        elif plan.kind == "remote":
                            payload = _fetch_remote(plan.remote_url)
                            plan.size_bytes = len(payload)
                            archive.writestr(arcname, payload, zipfile.ZIP_STORED)
                            plan.status = "exported"
                        elif plan.kind == "skipped":
                            plan.status = "skipped"
                        else:
                            plan.status = "missing"
                    except Exception as exc:  # one bad image must not kill the export
                        plan.status = "failed"
                        plan.note = f"{type(exc).__name__}: {exc}"[:300]
                        logger.warning("Product export could not add %s: %s", plan.source, exc)

                    if plan.status == "exported":
                        result.images_exported += 1
                        result.bytes_written += plan.size_bytes
                    elif plan.status != "skipped":
                        result.images_failed += 1
                        result.failures.append(f"{product.slug}: {plan.source} ({plan.note or plan.status})")

                    if buffer.pending() >= STREAM_CHUNK_BYTES:
                        yield buffer.take()

        workbook = build_workbook(products, plans_by_pk, options=options)
        archive.writestr(f"{root_name}/products.xlsx", workbook_bytes(workbook), zipfile.ZIP_DEFLATED)
        yield buffer.take()

        archive.writestr(f"{root_name}/README.txt", _readme_text(len(products), options), zipfile.ZIP_DEFLATED)
        if result.failures:
            archive.writestr(
                f"{root_name}/images-not-exported.txt",
                "These image references could not be written into the export:\n\n"
                + "\n".join(result.failures[:2000])
                + "\n",
                zipfile.ZIP_DEFLATED,
            )

    yield buffer.take()


def write_export_directory(
    products,
    plans_by_pk,
    destination: Path,
    *,
    include_images=True,
    result=None,
    options=None,
    progress=None,
):
    """Write the export as a real folder tree under ``destination``."""
    result = result if result is not None else ExportResult()
    result.products = len(products)
    destination.mkdir(parents=True, exist_ok=True)

    if include_images:
        for position, product in enumerate(products, start=1):
            for plan in plans_by_pk.get(product.pk, []):
                target = destination / plan.arcname
                try:
                    target.parent.mkdir(parents=True, exist_ok=True)
                    if plan.kind == "local" and plan.local_path and plan.local_path.is_file():
                        shutil.copy2(plan.local_path, target)
                        plan.size_bytes = target.stat().st_size
                        plan.status = "exported"
                    elif plan.kind == "remote":
                        payload = _fetch_remote(plan.remote_url)
                        target.write_bytes(payload)
                        plan.size_bytes = len(payload)
                        plan.status = "exported"
                    elif plan.kind == "skipped":
                        plan.status = "skipped"
                    else:
                        plan.status = "missing"
                except Exception as exc:
                    plan.status = "failed"
                    plan.note = f"{type(exc).__name__}: {exc}"[:300]
                    logger.warning("Product export could not copy %s: %s", plan.source, exc)

                if plan.status == "exported":
                    result.images_exported += 1
                    result.bytes_written += plan.size_bytes
                elif plan.status != "skipped":
                    result.images_failed += 1
                    result.failures.append(f"{product.slug}: {plan.source} ({plan.note or plan.status})")
            if progress:
                progress(position, len(products), product)

    workbook = build_workbook(products, plans_by_pk, options=options)
    (destination / "products.xlsx").write_bytes(workbook_bytes(workbook))
    (destination / "README.txt").write_text(_readme_text(len(products), options), encoding="utf-8")
    if result.failures:
        (destination / "images-not-exported.txt").write_text(
            "These image references could not be written into the export:\n\n"
            + "\n".join(result.failures[:2000])
            + "\n",
            encoding="utf-8",
        )
    return result


def zip_directory(source: Path, archive_path: Path):
    """Zip an already-written export folder, keeping its top folder inside."""
    with zipfile.ZipFile(archive_path, "w", allowZip64=True) as archive:
        for path in sorted(source.rglob("*")):
            if not path.is_file():
                continue
            arcname = f"{source.name}/{path.relative_to(source).as_posix()}"
            compression = zipfile.ZIP_DEFLATED if path.suffix.lower() in {".xlsx", ".txt", ".csv"} else zipfile.ZIP_STORED
            archive.write(path, arcname, compress_type=compression)
    return archive_path.stat().st_size


def human_bytes(size):
    value = float(size or 0)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{value:.1f} GB"


__all__ = [
    "ExportResult",
    "ImagePlan",
    "build_plans",
    "build_workbook",
    "export_base_name",
    "export_queryset",
    "human_bytes",
    "iter_export_zip",
    "plan_product_images",
    "workbook_bytes",
    "write_export_directory",
    "zip_directory",
]
